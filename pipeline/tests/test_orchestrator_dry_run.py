"""End-to-end dry-run test for the orchestrator.

This wires the full pipeline together using GIC_DRY_RUN=true so no LLM calls
happen. It verifies that the orchestrator can produce a populated Excel file
from an empty sources_dir using stub responses.
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook


def test_dry_run_end_to_end(tmp_path, monkeypatch, template_xlsx):
    if not template_xlsx.exists():
        pytest.skip(f"Template not present: {template_xlsx}")

    # Empty sources dir
    sources = tmp_path / "sources"
    sources.mkdir()

    output_dir = tmp_path / "out"
    monkeypatch.setenv("GIC_DRY_RUN", "true")

    from prefill.orchestrator import PipelineOrchestrator
    orch = PipelineOrchestrator.from_env(
        sources_dir=str(sources),
        template_xlsx=str(template_xlsx),
        output_dir=str(output_dir),
    )
    summary = orch.run()

    # Output files present
    assert Path(summary["app_facts"]).exists()
    assert Path(summary["responses"]).exists()
    assert Path(summary["output_xlsx"]).exists()
    assert Path(summary["audit_log"]).exists()

    # In dry-run mode, no LLM calls should have been made
    assert summary["llm_calls"] == 0

    # The populated workbook should have all 332 responses (all "Not Sure")
    wb = load_workbook(summary["output_xlsx"])
    ws = wb["03_Assessment"]
    populated = sum(1 for r in range(5, ws.max_row + 1) if ws.cell(r, 5).value)
    assert populated == 332

    # All responses should be "Not Sure" in dry-run
    not_sure = sum(1 for r in range(5, ws.max_row + 1) if ws.cell(r, 5).value == "Not Sure")
    assert not_sure == 332
