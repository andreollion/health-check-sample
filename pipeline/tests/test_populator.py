"""End-to-end test for the Excel populator using the real v0.9 template."""
from __future__ import annotations

import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook
from prefill.populator import populate_checklist
from prefill.skills_loader import load_health_check_skill


@pytest.fixture
def all_elements(health_skill_dir):
    skill = load_health_check_skill(health_skill_dir)
    return [
        {
            "element_id": el["element_id"],
            "qid": q["qid"],
            "level": el["level"],
            "response": "not_sure",
            "confidence": "low",
            "justification": f"Test justification for {el['element_id']}",
            "evidence_refs": ["a.b.c"],
        }
        for q in skill.all_questions
        for el in q["elements"]
    ]


def test_populator_writes_all_elements(tmp_path, template_xlsx, health_skill_dir, all_elements):
    if not template_xlsx.exists():
        pytest.skip(f"Template not present: {template_xlsx}")

    skill = load_health_check_skill(health_skill_dir)
    out = tmp_path / "populated.xlsx"
    populate_checklist(
        template_xlsx=template_xlsx,
        output_xlsx=out,
        responses=all_elements,
        health_skill=skill,
    )

    wb = load_workbook(str(out))
    ws = wb["03_Assessment"]

    # Count rows with a non-null Response value (col E)
    populated = 0
    for r in range(5, ws.max_row + 1):
        if ws.cell(r, 5).value:
            populated += 1
    assert populated == 332, f"expected 332 responses populated, got {populated}"


def test_populator_with_yes_response(tmp_path, template_xlsx, health_skill_dir):
    if not template_xlsx.exists():
        pytest.skip(f"Template not present: {template_xlsx}")

    skill = load_health_check_skill(health_skill_dir)
    responses = [{
        "element_id": "Q01-E01",
        "qid": "Q01",
        "level": "L1",
        "response": "yes",
        "confidence": "high",
        "justification": "Specifically tested",
        "evidence_refs": ["identity.sso.enabled"],
    }]
    out = tmp_path / "populated_one.xlsx"
    populate_checklist(
        template_xlsx=template_xlsx,
        output_xlsx=out,
        responses=responses,
        health_skill=skill,
    )

    wb = load_workbook(str(out))
    ws = wb["03_Assessment"]
    # Find the first element row (row 6 based on inspection)
    assert ws.cell(6, 5).value == "Yes"
    note = ws.cell(6, 6).value or ""
    assert "Specifically tested" in note
    assert "confidence: high" in note


def test_populator_app_profile(tmp_path, template_xlsx, health_skill_dir, all_elements):
    if not template_xlsx.exists():
        pytest.skip(f"Template not present: {template_xlsx}")

    skill = load_health_check_skill(health_skill_dir)
    app_facts = {
        "app": {
            "name": "payments-svc",
            "asset_id": "AST-12345",
            "asset_status": "Active",
            "scr_rating": "A",
            "classification": "Confidential",
        }
    }
    out = tmp_path / "populated_profile.xlsx"
    populate_checklist(
        template_xlsx=template_xlsx,
        output_xlsx=out,
        responses=all_elements,
        health_skill=skill,
        app_facts=app_facts,
    )

    wb = load_workbook(str(out))
    ws = wb["02_App_Profile"]
    label_to_val = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(1, ws.max_row + 1)}
    assert label_to_val.get("Application Name") == "payments-svc"
    assert label_to_val.get("Asset ID") == "AST-12345"
    assert label_to_val.get("Classification") == "Confidential"
