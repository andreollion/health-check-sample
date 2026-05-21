"""Writes responses into the v0.9 Checklist Excel template.

Row-matching strategy:
  - Read the '03_Assessment' sheet
  - Walk rows from header (row 4) onward
  - When QID column has a value (e.g. 'Q01'), that's the parent question row;
    record the current parent and reset element counter
  - When QID is empty but ParentQID is set, this is an element row; assign
    element_id = f"{parent_qid}-E{counter:02d}" and write the response

Response value mapping (canonical):
  - "yes"      → "Yes"
  - "no"       → "No"
  - "n/a"      → "N/A"
  - "not_sure" → "Not Sure"
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..skills_loader import HealthCheckSkill

logger = logging.getLogger(__name__)


RESPONSE_LABELS = {
    "yes": "Yes",
    "no": "No",
    "n/a": "N/A",
    "not_sure": "Not Sure",
}


# Column indices for the 03_Assessment sheet (1-based, openpyxl convention)
COL_QID = 1
COL_DIMENSION = 2
COL_SUBDOMAIN = 3
COL_TEXT = 4
COL_RESPONSE = 5
COL_NOTE = 6
COL_LEVEL = 10
COL_PARENT_QID = 11

HEADER_ROW = 4


def _build_response_index(responses: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """Index responses by element_id."""
    idx: dict[str, dict[str, Any]] = {}
    for r in responses:
        eid = r.get("element_id")
        if eid:
            idx[eid] = r
    return idx


def _format_note(r: dict[str, Any]) -> str:
    """Compose the 'Link or Brief Note' cell text."""
    bits: list[str] = []
    just = (r.get("justification") or "").strip()
    if just:
        bits.append(just)

    refs = r.get("evidence_refs") or []
    if refs:
        ref_str = ", ".join(refs[:5])
        bits.append(f"[evidence: {ref_str}]")

    conf = r.get("confidence")
    if conf:
        bits.append(f"[confidence: {conf}]")

    return " ".join(bits)


def populate_app_profile(ws, app_facts: dict[str, Any] | None) -> None:
    """Write app metadata into 02_App_Profile if available."""
    if not app_facts:
        return
    app = app_facts.get("app", {}) or {}
    identity = app_facts.get("identity", {}) or {}

    # Row-by-row mapping based on observed v0.9 structure
    mapping = {
        "Application Name": app.get("name"),
        "Asset ID": app.get("asset_id"),
        "Asset Status": app.get("asset_status"),
        "SCR Rating": app.get("scr_rating"),
        "Primary AppsOIC": app.get("apps_oic") or app.get("owner"),
        "Classification": app.get("classification") or app.get("data_classification"),
        "External Connectivity": app.get("external_connectivity"),
        "Product Type": app.get("product_type"),
    }
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if not label:
            continue
        if label in mapping and mapping[label] is not None:
            ws.cell(r, 2).value = mapping[label]


def populate_assessment(ws, responses: list[dict[str, Any]]) -> tuple[int, int, list[str]]:
    """Write Response + Note for every element row in 03_Assessment.

    Returns (written_count, missing_count, unmatched_eids).
    """
    response_index = _build_response_index(responses)

    written = 0
    missing = 0
    unmatched: list[str] = []

    current_parent: str | None = None
    counter = 0

    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        qid_val = ws.cell(r, COL_QID).value
        parent_val = ws.cell(r, COL_PARENT_QID).value
        text_val = ws.cell(r, COL_TEXT).value

        # Parent row: has QID, may have no ParentQID
        if qid_val and str(qid_val).strip().startswith("Q"):
            current_parent = str(qid_val).strip()
            counter = 0
            continue

        # Skip blank rows
        if not text_val and not parent_val:
            continue

        # Element row: ParentQID set, QID empty
        if parent_val and (not qid_val or str(qid_val).strip() == ""):
            parent_qid = str(parent_val).strip()
            if parent_qid != current_parent:
                # Shouldn't happen if the template is well-formed, but
                # adjust counter in case rows are out of order.
                current_parent = parent_qid
                counter = 0
            counter += 1
            element_id = f"{current_parent}-E{counter:02d}"

            resp = response_index.get(element_id)
            if not resp:
                missing += 1
                unmatched.append(element_id)
                continue

            ws.cell(r, COL_RESPONSE).value = RESPONSE_LABELS.get(
                resp.get("response", "not_sure"), "Not Sure"
            )
            ws.cell(r, COL_NOTE).value = _format_note(resp)
            written += 1

    return written, missing, unmatched


def populate_checklist(
    template_xlsx: Path,
    output_xlsx: Path,
    responses: list[dict[str, Any]],
    health_skill: HealthCheckSkill,
    app_facts: dict[str, Any] | None = None,
) -> Path:
    """End-to-end: load template, populate, save to output_xlsx. Returns output path."""
    template_xlsx = Path(template_xlsx)
    output_xlsx = Path(output_xlsx)
    if not template_xlsx.exists():
        raise FileNotFoundError(f"Template not found: {template_xlsx}")

    logger.info(f"loading template: {template_xlsx}")
    wb = load_workbook(str(template_xlsx))

    if "03_Assessment" not in wb.sheetnames:
        raise ValueError(f"Template missing '03_Assessment' sheet; sheets: {wb.sheetnames}")

    written, missing, unmatched = populate_assessment(wb["03_Assessment"], responses)
    logger.info(f"03_Assessment: wrote {written}, missing {missing}")
    if unmatched[:10]:
        logger.warning(f"first 10 unmatched element_ids: {unmatched[:10]}")

    if "02_App_Profile" in wb.sheetnames:
        populate_app_profile(wb["02_App_Profile"], app_facts)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_xlsx))
    logger.info(f"saved populated workbook: {output_xlsx}")
    return output_xlsx
